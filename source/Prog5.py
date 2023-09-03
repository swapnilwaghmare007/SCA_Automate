from requests_html import HTMLSession
import openpyxl

def are_values_remaining(sheet):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[13] is None:  # Check if column 14 (index 13) is empty
            return True
    return False

workbook = openpyxl.load_workbook('rescored_workbook.xlsx')
sheet = workbook.active
session = HTMLSession()

while are_values_remaining(sheet):
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[13] is None:  # Check if column 14 (index 13) is empty
            cvss_vector_full = row[15]
            cvss_vector = cvss_vector_full[9:]
            url = f"https://www.first.org/cvss/calculator/3.1#CVSS:3.1/{cvss_vector}"
            response = session.get(url)
            response.html.render()
            base_score_element = response.html.find('#environmentalMetricScore', first=True)

            if base_score_element:
                base_score_text = base_score_element.text
                if base_score_text:
                    base_score = float(base_score_text)
                    sheet.cell(row=index, column=14, value=base_score)
                    print(f"CVSS Vector: {cvss_vector_full} | Base Score: {base_score} (Updated)")
                else:
                    print(f"CVSS Vector: {cvss_vector_full} | Base score not found.")
            else:
                print(f"CVSS Vector: {cvss_vector_full} | Base score element not found.")

    # Save the workbook after each iteration
    workbook.save('rescored_workbook.xlsx')

print("\nNew Score Updated under NC CVSS v3.\n")
