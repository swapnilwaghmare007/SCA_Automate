import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook("rescored_workbook.xlsx")
sheet = workbook.active

# Define severity ranges and labels
severity_ranges = [(0.1, 3.9), (4.0, 6.9), (7.0, 8.9), (9.0, 10.0)]
severity_labels = ["Low", "Medium", "High", "Critical"]

# Update columns starting from the second row
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=14, max_col=14):
    value = row[0].value
   
    if value == "Not Applicable":
         continue      
    
    if value is not None:
        severity = None
        for i, (lower, upper) in enumerate(severity_ranges):
            if lower <= value <= upper:
                severity = severity_labels[i]
                break
        
    if severity is not None:
            sheet.cell(row=row[0].row, column=18, value=f"Rescored to ({severity})") # Update status column
            sheet.cell(row=row[0].row, column=5, value=severity)
                
    
         


print("\nNew Severities Updated\n.")
workbook.save("rescored_workbook.xlsx")


