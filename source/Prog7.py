import openpyxl

def make_list_vendorImages(db_vendorImages):
    # reading the database file
    data = db_vendorImages.read()
    # spliting data with (\n)
    data_into_list = data.split("\n")
    return data_into_list
def process_vendor_fix(source_file, vendorImage_list):
    
    # Assuming First sheet of the file
    source_sheet = source_file.active
    status='vendor fix'
    comment='requires ticket creation to vendor'
    for row_index, row in enumerate(source_sheet.iter_rows(min_row=2, values_only=True), start=2):
        Artifact = row[20]  # 21st column
        # print(Artifact)
        
        # print(Status)
        vendorImage = Artifact.split('/', 1)
        if vendorImage[0] in vendorImage_list:
            # print(True)
            source_sheet.cell(row=row_index, column=18, value=status)
            source_sheet.cell(row=row_index, column=19, value=comment)

    source_file.save("rescored_workbook.xlsx")
    print("\nVendor Fix Check Completed\n'")
    
    source_file.close()

if __name__ == "__main__":
    db_file = open("dataset/VendorImages.txt", "r")
    source_file = openpyxl.load_workbook('rescored_workbook.xlsx')
    vendorImage_list = make_list_vendorImages(db_file)
    process_vendor_fix(source_file, vendorImage_list)