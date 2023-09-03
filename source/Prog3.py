import openpyxl
import json
def read_config_values(config_file):
    with open(config_file, 'r') as f:
        data = f.read()
    rules = json.loads(data)
    return rules

def modify_vector(ogVector, mVector):
    if len(ogVector) >= 44:
        original_vector = str(ogVector)
        new_vector = original_vector[:44]+ mVector  # Use the first vector from the config file
    else:
        new_vector = ogVector
    # print(new_vector)
    return new_vector

def create_microservice_dict(database_file_path):
    db_file = openpyxl.load_workbook(database_file_path)
    db_sheet = db_file.active
    db_microservices = {}
    for row in db_sheet.iter_rows(min_row=2, values_only=True):
        microservice = row[0]  # Assuming Microservice is in the first column (index 0)
        trimmed_microservice = microservice.strip()  # Trim leading and trailing spaces
        db_microservices[trimmed_microservice] = True
    return db_microservices

def process_xlsx_file(file_path, config_values, db_microservices):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        package_type = row[11]  # 12th column, 0-indexed
        # print(package_type)
        impact_path = row[9]  # 10th column
        microservice = row[2]  # 3rd column
        # print(microservice)
        vector = row[14]  # 16th column
        
        if ((package_type in config_values.keys()) and (microservice in db_microservices) and ('/diagtools' not in impact_path)):
            index = list(config_values.keys()).index(package_type)
            if (index == 0):
                continue
            else:
                new_vector = modify_vector(vector, config_values.get(package_type))
        elif ((package_type in config_values.keys()) and ('/diagtools' in impact_path)):
            new_vector = modify_vector(vector, config_values.get(package_type))
        else:
            continue

        sheet.cell(row=row_index, column=16, value=new_vector)

    workbook.save('rescored_workbook.xlsx')


if __name__ == "__main__":
    file_path = "rescored_workbook.xlsx"
    config_file_path = "ruleset/PackageRuleSet.txt"
    db_file_path = "dataset/baseImageDataset.xlsx"
    config_values = read_config_values(config_file_path)
    db_microservices = create_microservice_dict(db_file_path)
    process_xlsx_file(file_path, config_values, db_microservices)