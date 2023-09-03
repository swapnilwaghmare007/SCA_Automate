import pandas as pd
from openpyxl import load_workbook

def concatenate_and_compare(file1_path, test_path):
    # Read the Excel files into pandas dataframes
    file1_df = pd.read_excel(file1_path)
    test_df = pd.read_excel(test_path)

    # Concatenate columns for both dataframes
    test_concatenated = test_df.iloc[:, 0] + '.' + test_df.iloc[:, 2]
    file1_concatenated = file1_df.iloc[:, 1] + '.' + file1_df.iloc[:, 2]

    # Check for matching records and update values
    for i, test_value in enumerate(test_concatenated):
        matching_index = file1_concatenated[file1_concatenated == test_value].index
        if not matching_index.empty:
            matching_row_index = matching_index[0]
            matching_row = file1_df.iloc[matching_row_index]

            test_13th_column_value = test_df.loc[i, test_df.columns[14]]  # Assuming the 13th column index is 12

            updated_value = str(test_13th_column_value)  # Start with 13th column value

            if pd.notna(matching_row[3]):  # Check if the 4th column value is not None or NaN
                updated_value += '/' + 'M' + str(matching_row[3])

            # Append values from 5th, 6th, and 7th columns
            for col_index in range(4, 7):
                if pd.notna(matching_row[col_index]):
                    updated_value += '/' + 'M' + str(matching_row[col_index])

            workbook = load_workbook(test_path)
            worksheet = workbook.active
            worksheet.cell(row=i + 2, column=16, value=updated_value)
            workbook.save(test_path)
            print(f"Updated value in row {i + 2}")
        else:
            print(f"No match found for row {i + 2}")
    
    print("\nVector Modification completed for Microservices\n")

if __name__ == "__main__":
    file1_path = 'dataset/microservices_dataset.xlsx'
    test_path = 'rescored_workbook.xlsx'
    concatenate_and_compare(file1_path, test_path)
