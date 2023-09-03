import pandas as pd
import openpyxl
from openpyxl.styles import Font
import json

# Read the Excel file
file_path = 'workbook.xlsx'
output_file_path = 'rescored_workbook.xlsx'
df = pd.read_excel(file_path)

# Read deletion rules from JSON file
with open('ruleset/InitialDeletionRules.txt', 'r') as rules_file:
    deletion_rules = json.load(rules_file)

# Apply deletion rules
for column, values_to_remove in deletion_rules.items():
    if column == "Fixed Versions" and values_to_remove == "[]":
        df = df[df[column] != values_to_remove]
    elif isinstance(values_to_remove, list):
        df = df[~df[column].isin(values_to_remove)]
    elif values_to_remove is None:
        df = df[df[column].isnull()]

# Additional filter
df = df[df['CVSS v3'] >= 6.9]

# Additional Insertion of Columns
df.insert(4, 'NC Severity', None)
df.insert(13, 'NC CVSS v3', None)
df.insert(15, 'NC CVSS Vector', None)
# Apply a specific rule for the "Microservice" column
microservice_value = deletion_rules.get("Microservice")
if microservice_value is not None:
    df = df[df["Microservice"] != microservice_value]  # Use != to remove the specified value

# Create a new Excel file using openpyxl
wb = openpyxl.Workbook()
ws = wb.active

# Apply bold font to the first row (column names)
font = Font(bold=True)
for c_idx, column_name in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=c_idx, value=column_name)
    cell.font = font

# Write the DataFrame into the Excel sheet, starting from the second row
for r_idx, row in enumerate(df.iterrows(), start=2):
    for c_idx, value in enumerate(row[1], start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook
wb.save(output_file_path)


